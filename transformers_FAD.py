###################
import numpy as np
import pandas as pd
import tensorflow as tf
import keras
import matplotlib.pyplot as plt
import os
import glob
import warnings
import math
import optuna
import logging
import sys
import gc
import socket

from tensorflow.keras.layers import Input, Dense, LayerNormalization, Dropout, MultiHeadAttention, GlobalAveragePooling1D
from tensorflow.keras.models import Model
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import MinMaxScaler
from scipy.stats import linregress
from keras.models import model_from_json
from keras.models import Sequential, Model
from keras.layers import LSTM, Dropout, Bidirectional, Dense, Input
from keras import backend as K
from tensorflow.keras import regularizers
from optuna.integration import KerasPruningCallback
from optuna.trial import TrialState
from openpyxl import load_workbook
from tensorflow.keras.utils import plot_model
warnings.filterwarnings("ignore")
# os.chdir('/content/drive/MyDrive/Doutorado/Testes_corretos')
hostname = socket.gethostname()

if hostname == "DESKTOP-7TSHE39":
    path = 'G:\Meu Drive\Doutorado\Testes_corretos'
    from datetime import datetime
    data_atual = datetime.now().strftime('%Y-%m-%d_%H-%M')
    caminho_novo = 'G:\\Meu Drive\\Doutorado\\Códigos\\transformers_FAD\\Resultados\\' + data_atual
    planilha_geral = 'G:\\Meu Drive\\Doutorado\\Códigos\\transformers_FAD\\geral.xlsx'
############

global look_back
look_back = 2             # number of time steps used to predict
ni=7                       # number of input variables
no=1                       # number of output variables
test_split_size = 0.2
filter_size = 60           # "moving-average" in seconds

# serialize model to JSON
# def save_model(model):
#   path_save = '/content/drive/MyDrive/Doutorado/Códigos/Transformers/Models/look_back_{look_back}_no_{no}'.format(look_back = str(look_back), no = str(no))
#   if not os.path.exists(path_save):
#     os.mkdir(path_save)
#   os.chdir(path_save)
#   model_json = model.to_json()
#   with open("model.json", "w") as json_file:
#       json_file.write(model_json)
#   # serialize weights to HDF5
#   model.save_weights("model.h5")
#   !cp model.h5 path_save
#   !cp model.json path_save
#   print("Saved model to disk")
#   return
  # # Load RNN

# def load_LSTM_model():
#   json_file = open('/content/drive/MyDrive/Doutorado/Códigos/Transformers/Models/model.json', 'r')
#   loaded_model_json = json_file.read()
#   json_file.close()
#   model = model_from_json(loaded_model_json)
#   # load weights into new model
#   model.load_weights("/content/drive/MyDrive/Doutorado/Códigos/Bayesian/model.h5")
#   print("Loaded model from disk")
#   return model

def normalizacao(x, maximos, minimos):
  dados_norm = 1-2*((maximos-x)/(maximos-minimos))
  return dados_norm

def denormalizacao(x, maximos, minimos):
  dados_denorm = maximos-((1-x)*(maximos-minimos))/2
  return dados_denorm

def ytest(Dy, fstep): # Dt must be a np array
  max_index = len(Dy)
  aux_array = np.empty([max_index-fstep, fstep])
  i = 0
  for i in range(max_index-fstep):
    aux_array[i, :] = Dy[i:i+fstep]
    i += 1
  return aux_array

# Creating a data structure with look_back timesteps and k outputs
def create_dataset(training_set_scaled, look_back):
    x_train, y_train = [], []
    for i in range(look_back, len(training_set_scaled)):
        x_train.append(training_set_scaled[i - look_back:i, 0:ni])
        y_train.append(training_set_scaled[i, ni:ni+no])
    return np.array(x_train).astype('float32'), np.array(y_train).astype('float32')

def criar_dataset(path):
  csv_files = glob.glob(os.path.join(path, "*.xlsx"))
  dataset = []
  first_loop = 0
  for file_name in csv_files:
      df = pd.read_excel(file_name, na_filter = False).drop(columns=['Time'])
      i = 0
      mean_values = []
      while i <= df.shape[0]:
        a = df[i:i+filter_size].mean()
        a.values.tolist()
        mean_values.append(a)
        i += filter_size
      m = pd.DataFrame(mean_values)
      outlet_turb = np.array(m['Outlet turbidity'])
      input_var = m[['Inlet Turbidity (NTU)', 'Inflow (L/min)', 'Recicle flow (L/min)', 'Pressure (bar)', 'Temperature (ºC)', 'Level (cm)', 'Outlet turbidity']]
      kDy_matrix = ytest(outlet_turb[look_back:len(outlet_turb)+1], no)
      dados_real = np.array(input_var)
      data = np.hstack([dados_real[0:len(dados_real)-look_back-no, :], kDy_matrix])
      x_d, y_d = create_dataset(data, look_back)
      if first_loop == 0:
        x_f = x_d
        y_f = y_d
        first_loop = 1
      else:
        x_f = np.append(x_f, x_d, axis=0)
        y_f = np.append(y_f, y_d, axis=0)
  return x_f, y_f

def norm_dataset(x_f, y_f):
  x_maximos = x_f.max(axis=0)
  x_minimos = x_f.min(axis=0)
  x_norm = normalizacao(x_f, x_maximos, x_minimos)
  x_train, x_test = train_test_split(x_norm, test_size=test_split_size, shuffle=False)

  y_maximos = y_f.max(axis=0)
  y_minimos = y_f.min(axis=0)
  y_norm = normalizacao(y_f, y_maximos, y_minimos)
  y_train, y_test = train_test_split(y_norm, test_size=test_split_size, shuffle=False)

  return x_train, x_test, y_train, y_test, x_maximos, x_minimos, y_maximos, y_minimos

x_f, y_f = criar_dataset(path)
global x_train, x_test, y_train, y_test
x_train, x_test, y_train, y_test, x_maximos, x_minimos, y_maximos, y_minimos = norm_dataset(x_f, y_f)

# Definição do modelo neural
def transformer_encoder(inputs, L1, L2, key_dim, num_heads, ff_dim, dropout):
    # Camada de multihead attention
    attention = MultiHeadAttention(key_dim=key_dim, num_heads=num_heads, dropout=dropout, kernel_regularizer=regularizers.L1L2(l1=L1, l2=L2))(inputs, inputs)
    attention = Dropout(dropout)(attention)
    attention = LayerNormalization(epsilon=1e-6)(attention)
    attention = attention + inputs

     # Camada feedforward
    outputs = Dense(ff_dim, activation='relu')(attention)
    outputs = Dropout(dropout)(outputs)
    outputs = Dense(inputs.shape[-1])(outputs)
    outputs = LayerNormalization(epsilon=1e-6)(outputs)
    return outputs + attention
def create_model(trial):
  L1 = trial.suggest_float("L1", 1e-5, 1e-1, log=True)
  L2 = trial.suggest_float("L2", 1e-5, 1e-1, log=True)
  num_heads = trial.suggest_int("num_heads", 2, 10)
  key_dim = trial.suggest_int("key_dim", 10, 100)
  ff_dim = trial.suggest_int("ff_dim", 1, 20)
  dropout_rate = trial.suggest_float("dropout_rate", 0.1, 0.5)
  lr_sug = trial.suggest_float('init_lr', 0.0000001, 0.01)
  K.clear_session()
  # tf.keras.backend.clear_session()

# Definindo o modelo
  input_layer = Input(shape=(look_back, ni))
  encoder = transformer_encoder(input_layer, L1, L2, key_dim= key_dim, num_heads=num_heads, ff_dim=ff_dim, dropout=dropout_rate)
  #encoder = transformer_encoder(encoder, L1, L2, key_dim = key_dim, num_heads=num_heads, ff_dim=ff_dim, dropout=dropout_rate)

  # Agregando a dimensão temporal usando Global Average Pooling
  pooled_output = GlobalAveragePooling1D()(encoder)

  # Camada de predição final
  output_layer = Dense(1)(pooled_output)

  modelo_transformer = Model(inputs=input_layer, outputs=output_layer)

  opt = keras.optimizers.Adam(learning_rate=lr_sug)

  modelo_transformer.compile(optimizer=opt, loss='mean_squared_error', metrics=['mse'])
  return modelo_transformer

def objective(trial):
  # Clear clutter from previous session graphs.
  es = tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=10)
  modelo_transformer = create_model(trial)
  modelo_transformer.fit(x_train, y_train, batch_size=32*5, callbacks=[KerasPruningCallback(trial, "val_loss"), es],
        epochs=50, validation_data=(x_test, y_test), verbose=0)
  score = modelo_transformer.evaluate(x_test, y_test, verbose=1)
  print(gc.collect())
  return score[1]

study = optuna.create_study(direction="minimize", pruner=optuna.pruners.SuccessiveHalvingPruner())
# study.optimize(objective, n_trials=250)
study.optimize(objective, n_trials=1000, gc_after_trial=True, callbacks=[lambda study, trial: gc.collect()])
pruned_trials = study.get_trials(deepcopy=False, states=[TrialState.PRUNED])
complete_trials = study.get_trials(deepcopy=False, states=[TrialState.COMPLETE])
print("Study statistics: ")
print("  Number of finished trials: ", len(study.trials))
print("  Number of pruned trials: ", len(pruned_trials))
print("  Number of complete trials: ", len(complete_trials))

print("Best trial:")
trial = study.best_trial

print("  Value: ", trial.value)

print("  Params: ")
for key, value in trial.params.items():
        print("    {}: {}".format(key, value))

def plotar_graficos_treinamento():
    N_PLOTTED_TRIALS = 10
    trials = study.trials
    plotted_trials = sorted(trials, key=lambda t: t.value)[:N_PLOTTED_TRIALS]
    plotted_study = optuna.create_study()
    for trial in plotted_trials:
        plotted_study.add_trial(trial)
    opt_history = optuna.visualization.plot_optimization_history(plotted_study)
    fig = optuna.visualization.plot_optimization_history(study)
    fig.write_html("optimization_history.html")
    fig2 = optuna.visualization.plot_intermediate_values(plotted_study)
    fig2.write_html("intermediate_values.html")
    fig3 = optuna.visualization.plot_contour(study)
    fig3.write_html("contour.html")
    importances = optuna.importance.get_param_importances(study)
    fig5 = optuna.visualization.plot_rank(study)
    fig5.write_html("rank.html")
    fig6 = optuna.visualization.plot_param_importances(study)
    fig6.write_html("importance.html")
    return

best_params = study.best_params
L1 = best_params['L1']
L2 = best_params['L2']
num_heads = best_params['num_heads']
key_dim = best_params['key_dim']
ff_dim = best_params['ff_dim']
dropout_rate = best_params['dropout_rate']
lr_sug = best_params['init_lr']

def final_model(key_dim, L1, L2, num_heads, ff_dim, dropout_rate, lr_sug):
# Definição do modelo neural
  def transformer_encoder(inputs, L1, L2, key_dim, num_heads, ff_dim, dropout):
    # Camada de multihead attention
    attention = MultiHeadAttention(key_dim=key_dim, num_heads=num_heads, dropout=dropout, kernel_regularizer=regularizers.L1L2(l1=L1, l2=L2))(inputs, inputs)
    attention = Dropout(dropout)(attention)
    attention = LayerNormalization(epsilon=1e-6)(attention)
    attention = attention + inputs

     # Camada feedforward
    outputs = Dense(ff_dim, activation='relu')(attention)
    outputs = Dropout(dropout)(outputs)
    outputs = Dense(inputs.shape[-1])(outputs)
    outputs = LayerNormalization(epsilon=1e-6)(outputs)
    return outputs + attention

# Definindo o modelo
  input_layer = Input(shape=(look_back, ni))
  encoder = transformer_encoder(input_layer, L1, L2, key_dim = key_dim, num_heads=num_heads, ff_dim=ff_dim, dropout=dropout_rate)
  #encoder = transformer_encoder(encoder, L1, L2, key_dim = key_dim, num_heads=num_heads, ff_dim=ff_dim, dropout=dropout_rate)

  # Agregando a dimensão temporal usando Global Average Pooling
  pooled_output = GlobalAveragePooling1D()(encoder)

  # Camada de predição final
  output_layer = Dense(no)(pooled_output)

  modelo_transformer = Model(inputs=input_layer, outputs=output_layer)

  opt = keras.optimizers.Adam(learning_rate=lr_sug)

  modelo_transformer.compile(optimizer=opt, loss='mean_squared_error', metrics=['mse'])
  es = tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=10)
  modelo_transformer = create_model(trial)
  modelo_transformer.fit(x_train, y_train, batch_size=64, callbacks=[KerasPruningCallback(trial, "val_loss"), es],
        epochs=50, validation_data=(x_test, y_test), verbose=0)
  return modelo_transformer

model = final_model(key_dim, L1, L2, num_heads, ff_dim, dropout_rate, lr_sug)
def save_model(model):
    model_json = model.to_json()
    with open("model_transformer.json", "w") as json_file:
        json_file.write(model_json)
    # serialize weights to HDF5
    model.save_weights("model_transformer.h5")
    print("Saved model to disk")
    return
os.mkdir(caminho_novo)
os.chdir(caminho_novo)
save_model(model)
plotar_graficos_treinamento()

train_predict = model.predict(x_train)
test_predict = model.predict(x_test)
train_score = mean_squared_error(y_train, train_predict)
test_score = mean_squared_error(y_test, test_predict)
y_denorm_test = denormalizacao(y_test, y_maximos, y_minimos)
ypred_denorm_test = denormalizacao(test_predict, y_maximos, y_minimos)
y_denorm_train = denormalizacao(y_train, y_maximos, y_minimos)
ypred_denorm_train = denormalizacao(train_predict, y_maximos, y_minimos)

line_width=2; marker_size=20; size=45
plt.rc('font', size=size)          # controls default text sizes
plt.rc('axes', titlesize=size)     # fontsize of the axes title
plt.rc('axes', labelsize=size)     # fontsize of the x and y labels
plt.rc('xtick', labelsize=size)    # fontsize of the tick labels
plt.rc('ytick', labelsize=size)    # fontsize of the tick labels
plt.rc('legend', fontsize=size)    # legend fontsize
plt.rc('figure', titlesize=size)   # fontsize of the figure title
plt.rcParams['axes.linewidth'] = 2  # set the value globally
plt.rcParams['font.family'] = 'serif'
plt.rcParams['font.serif'] = ['Times New Roman'] + plt.rcParams['font.serif']

x = np.linspace(0, len(y_denorm_test)-1, len(y_denorm_test))
# test, axs = plt.subplots(math.ceil(no/2), 2, figsize=(15, 15))
x = np.linspace(0, (len(y_denorm_test)-1)*filter_size/60, len(y_denorm_test))
test = plt.figure(figsize=(20, 15))
plt.plot(x, ypred_denorm_test, color='#0000ff', label = "Prediction")
plt.plot(x, y_denorm_test, color='#ff0000', label = "Data")
plt.legend()
plt.ylabel('Turbidity (NTU)')
plt.xlabel('Time (min)')
plt.tight_layout()
plt.savefig('test.png', dpi=600)

x = np.linspace(0, len(y_denorm_train)-1, len(y_denorm_train))
# test, axs = plt.subplots(math.ceil(no/2), 2, figsize=(15, 15))
x = np.linspace(0, (len(y_denorm_train)-1)*filter_size/60, len(y_denorm_train))
train = plt.figure(figsize=(20, 15))
plt.plot(x, ypred_denorm_train, color='#0000ff', label = "Prediction")
plt.plot(x, y_denorm_train, color='#ff0000', label = "Data")
plt.legend()
plt.ylabel('Turbidity (NTU)')
plt.xlabel('Time (min)')
plt.tight_layout()
plt.savefig('train.png', dpi=600)

# Caminho do arquivo Excel existente
wb = load_workbook(planilha_geral)
planilha = wb.active
proxima_linha = planilha.max_row + 1
novos_dados = [data_atual, look_back, key_dim, L1, L2, num_heads, ff_dim, dropout_rate, lr_sug, train_score, test_score]
for coluna, valor in enumerate(novos_dados, start=1):
    planilha.cell(row=proxima_linha, column=coluna, value=valor)
wb.save(planilha_geral)

df1 = pd.DataFrame(ypred_denorm_train)
df1.to_excel('dados_calculados_treinamento.xlsx', sheet_name='Treinamento')
df2 = pd.DataFrame(ypred_denorm_test)
df2.to_excel('dados_calculados_teste.xlsx', sheet_name='Teste')

with open(caminho_novo + '\\model_summary.txt', "w") as f:
    # Use a função summary() do modelo com o argumento print_fn para redirecionar a saída
    model.summary(print_fn=lambda x: f.write(x + "\n"))

